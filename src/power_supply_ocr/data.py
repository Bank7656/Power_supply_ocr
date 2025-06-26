import os
from openpyxl import Workbook
import pandas as pd


OUTPUT_DIR_NAME = "output"
TITLE_NAME = "power_supply_data"

current_row_index = 2

def create_output_dir() -> None:
    try:
        os.mkdir(OUTPUT_DIR_NAME)
    except FileExistsError:
        print(f"Directory {OUTPUT_DIR_NAME} already exists.")
    except OSError as e:
        print(f"Error creating directory: {e}")
    return


def create_excel():

    excel = Workbook()
    sheet = excel.active
    sheet.title = TITLE_NAME
    # Add headers
    sheet['A1'] = "Voltage (V)"
    sheet['B1'] = "Current (A)"
    return excel, sheet

def update_excel(sheet, value):
    voltage = value[0]
    current = value[1]
    global current_row_index
    try:
        sheet.cell(row=current_row_index, column=1, value=voltage)
        sheet.cell(row=current_row_index, column=2, value=current)
    except ValueError as e:
        print(f"WARNING: Skipping data pair due to conversion error at row {current_row_index}: {e}")
    except Exception as e:
        print(f"An unexpected error occurred while writing data: {e}")
    current_row_index += 1
    return

def save_excel(excel, filename):
    excel_file = OUTPUT_DIR_NAME + "/" + filename + ".xlsx"
    try:
        excel.save(excel_file)
    except PermissionError:
        print("Error: Permission denied. Please ensure the file is not open and you have write permissions for the location.")
    return
    
